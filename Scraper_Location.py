import logging
import time

from copy import copy
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException

ubci = []
uberlocsl = load_workbook(filename='/home/sameer/Uber_Cities.xlsx',use_iterators=True)
locsl = uberlocsl['Sheet1']
for row in locsl.iter_rows():
	for cell in row:
		ubci.append(cell.value)

wb = Workbook()
ws = wb.active
ws.title = "Uber Demographics"

driver = webdriver.Chrome('/home/sameer/bin/chromedriver')

driver.get('http://www.linkedin.com/');

login_box = driver.find_element_by_name('session_key')
password_box = driver.find_element_by_name('session_password')
login_box.send_keys('<INSERT LINKEDIN USERNAME HERE>')
password_box.send_keys('<INSERT LINKEDIN PASSWORD HERE>')
password_box.submit()

driver.get('https://www.linkedin.com/company/1815218')
driver.find_element_by_link_text('See all').click()

tot = 0
i = 1
for city in ubci:
	ws.cell(row = i, column = 1).value = city
	loc = driver.find_element_by_id('facet-G')
	loc.find_element_by_class_name('add-facet-button').click()
	time.sleep(3)
	add_box = loc.find_element_by_class_name('facet-typeahead')
	add_box.send_keys(city)
	time.sleep(3)
	try:
		print driver.find_element_by_class_name('item-headline').text,
		ws.cell(row = i, column = 2).value = driver.find_element_by_class_name('item-headline').text
		driver.find_element_by_class_name('item-headline').click()
		time.sleep(5)
		res = driver.find_element_by_id('results_count')
		print res.find_element_by_tag_name('strong').text
		ws.cell(row = i, column = 3).value = int(res.find_element_by_tag_name('strong').text)
		tot += int(res.find_element_by_tag_name('strong').text)
		print tot
	except:
		#the city we passed is not recognized by linkedin
		print 'No Results'
		ws.cell(row = i, column = 2).value = 'No Results'
		ws.cell(row = i, column = 3).value = 'No Results'
	driver.get('https://www.linkedin.com/vsearch/p?f_CC=1815218&trk=rr_connectedness')
	i += 1
	
wb.save('Uber Demographics.xlsx')
