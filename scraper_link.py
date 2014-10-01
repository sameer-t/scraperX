#!/usr/bin/env python
# -*- coding: utf-8 -*-

from selenium import webdriver
from openpyxl import load_workbook
import time

# Loading list of major loactions where Uber employees are loacted
ubci = []
uberlocsl = load_workbook(filename='/home/sameer/linkedin-scraper/Uber_Cities.xlsx',use_iterators=True)
locsl = uberlocsl['Sheet1']
for row in locsl.iter_rows():
    for cell in row:
        ubci.append(cell.value)

# Load driver.
driver = webdriver.Chrome('/usr/bin/chromedriver')

# Go to LinkedIn.
driver.get('http://www.linkedin.com/');

# Login
login_box = driver.find_element_by_name('session_key')
password_box = driver.find_element_by_name('session_password')
login_box.send_keys('<INSERT LINKEDIN USERNAME HERE>')
password_box.send_keys('<INSERT LINKEDIN PASSWORD HERE>')
password_box.submit()

# Navigate to Uber LinkedIn Page
driver.get('https://www.linkedin.com/company/1815218')
time.sleep(3)
driver.find_element_by_link_text('See all').click()

# Get LinkedIn profile links citywise
allusers = []
for city in ubci:
    time.sleep(5)
    loc = driver.find_element_by_id('facet-G')
    loc.find_element_by_class_name('add-facet-button').click()
    time.sleep(3)
    add_box = loc.find_element_by_class_name('facet-typeahead')
    add_box.send_keys(city)
    time.sleep(3)
    driver.find_element_by_class_name('item-headline').click()
    time.sleep(5)
    try:
        res = driver.find_element_by_id('results_count')
        res = res.find_element_by_tag_name('strong').text
        print city + ' has ' + res + ' employees'
        users = []
        while len(users) < users:
            users += [element.get_attribute('href') for element in driver.find_elements_by_class_name('title')]
            print len(users)
            driver.get(driver.find_element_by_link_text('Next >').get_attribute('href'))
    except:
        #the city we passed is not recognized by linkedin
        if int(res) > len(users):
            print 'No Results for ' + city
    driver.get('https://www.linkedin.com/vsearch/p?f_CC=1815218&trk=rr_connectedness')
    allusers += users

# Saving in a file to be used by the scraper_html script
links = open('links_uber.txt', 'w')
links.write(str(allusers))
links.close()

driver.quit()

print 'Link scraping completed ( ͡° ͜ʖ ͡°)'