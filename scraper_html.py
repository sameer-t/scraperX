#!/usr/bin/env python
# -*- coding: utf-8 -*-

from selenium import webdriver
from openpyxl import Workbook
import time

# Load driver.
driver = webdriver.Chrome('/usr/bin/chromedriver')

# Go to LinkedIn.
driver.get('http://www.linkedin.com/');

# Login
login_box = driver.find_element_by_name('session_key')
password_box = driver.find_element_by_name('session_password')
login_box.send_keys('<INSERT USERNAME HERE>')
password_box.send_keys('<INSERT PASSWORD HERE>')
password_box.submit()

# Load the links created using the scraper_link script
links = open('links_uber', 'r')
users = eval(links.read())
print len(users)

# Create a workbook with name and profile url for later use
wb = Workbook()
ws = wb.active

errorlinks = []

# Start saving all the profiles in html format in the local system. 
# If number of users is more than 500, consider using multiple premium accounts for scraping
userno = 1
for ulink in users:
    try:
        if str(type(ulink)) == "<type 'unicode'>":
            driver.get(str(ulink))
            time.sleep(5)
            name = driver.find_element_by_class_name('full-name').text
            print str(userno) + ' ' + name
            page_html = driver.page_source.encode('utf-8')
            user = open(str(userno)+'.html','w')
            user.write(page_html)
            user.close()
            ws.cell(row = userno, column = 1).value = name.decode('utf-8','ignore')
            ws.cell(row = userno, column = 2).value = str(ulink)
            userno +=1
        else:
            print 'Not a link (ノ ゜Д゜)ノ ︵ ┻━┻'
    except:
        errorlinks.append(ulink)
        print 'Some exception (-_-メ)'

if len(errorlinks) != 0:
    fails = open('uber_failed_links', 'w')
    fails.write(str(errorlinks))
    fails.close()

wb.save('uber_lookup.xlsx')

driver.quit()

print 'HTML scraping completed ( ͡° ͜ʖ ͡°)'