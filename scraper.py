#!/usr/bin/env python
# -*- coding: utf-8 -*-

import logging
import time
import csv

from copy import copy
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException

def trim (des):
	desc = des.encode('utf8')
	toremove = ['•', '-','\n1.','\n2.','\n3.','\n4.','\n5.','\n6.','\n7.','\n8.','\n9.','\n10.','\n11.','\n12.', '\n13.', '\n14.', '\n15', '\n\n']
	for symbol in toremove:
		if symbol in desc:
			desc = desc.replace(symbol,'') 
	return desc

ubci = []
uberlocsl = load_workbook(filename='/home/sameer/Uber_Cities2.xlsx',use_iterators=True)
locsl = uberlocsl['Sheet1']
for row in locsl.iter_rows():
	for cell in row:
		ubci.append(cell.value)

# Load driver.
driver = webdriver.Chrome('/home/sameer/bin/chromedriver')

# Go to LinkedIn.
driver.get('http://www.linkedin.com/');

# Login
login_box = driver.find_element_by_name('session_key')
password_box = driver.find_element_by_name('session_password')
login_box.send_keys('<INSERT LINKEDIN USERNAME HERE>')
password_box.send_keys('<INSERT LINKEDIN PASSWORD HERE>')
password_box.submit()

# Go to Uber page.
driver.get('https://www.linkedin.com/company/1815218')

# Get to list page.
driver.find_element_by_link_text('See all').click()

# Start scraping.

allusers = []

for city in ubci:
	time.sleep(5)
	loc = driver.find_element_by_id('facet-G')
	loc.find_element_by_class_name('add-facet-button').click()
	time.sleep(3)
	add_box = loc.find_element_by_class_name('facet-typeahead')
	add_box.send_keys(city)
	time.sleep(3)
	driver.find_element_by_class_name('item-headline').click()
	time.sleep(5)
	try:
		users = []
		while len(users) < users:
			users += [element.get_attribute('href') for element in driver.find_elements_by_class_name('title')]
			print len(users)
			driver.get(driver.find_element_by_link_text('Next >').get_attribute('href'))
	except Exception:
		#print users
		links = open('links', 'w')
		links.write(str(users))
		print 'Acquired links to user profiles for ' + city
	allusers += users
	driver.get('https://www.linkedin.com/vsearch/p?f_CC=1815218&trk=rr_connectedness')

alllinks = open('alllinks', 'w')
alllinks.write(str(allusers))

alllinks = open('./alllinks', 'r')
users = eval(alllinks.read())
print len(users)

alluserdat = []
userno = 1 
try:
	for link in users:
		if str(type(link)) == "<type 'unicode'>":
			driver.get(str(link))
			print userno
			userdat = []
			data = ''
			try:
				#Name
				try:
					userdat.append(str(userno) + '. ' + driver.find_element_by_class_name('full-name').text + '\n')
					print 'Name Done'
				except:
					userdat.append('No Name')
					print 'No Name'			
				#Title
				try:
					data = data + trim(driver.find_element_by_class_name('title').text) + '\n'
					print 'Title Done'
				except:
					print 'No Title' 
				#Summary						
				try:
					wrapper1 = driver.find_element_by_class_name('summary')
					data = data + trim(wrapper1.find_element_by_class_name('description').text) + '\n'
				except NoSuchElementException:
					print 'No Summary'
				print 'Summary Done'
				#Skills
				try:
					wrapper3 = driver.find_elements_by_class_name('endorse-item-name')
					for v in range(len(wrapper3)):
						ski = wrapper3[v].text.replace('...','')
						if ski != '':
							data = data + trim(ski) + '\n'
					print 'Skills Done'
				except:
					print 'No Skills'
				#Education
				try:
					wrapper4 = driver.find_element_by_id('background-education')
					er = wrapper4.find_elements_by_class_name('editable-item')
					for i in range(len(er)):
						data = data + trim(er[i].text) + '\n'
					print 'Edu Done'
				except:
					print 'No Education'
				#Additional Info - Interests
				try:
					wrapper5 = driver.find_element_by_class_name('interests-listing')
					data = data + trim(wrapper5.text) + '\n'
				except NoSuchElementException:
					print 'No interests'
				print 'Int Done'
				#Experience
				try:
					wrapper2 = driver.find_element_by_id('background-experience')
					qw = wrapper2.find_elements_by_class_name('editable-item')
					for k in range(len(qw)):
						data = data + trim(qw[k].text) + '\n'
					print 'Exp Done'
				except:
					print 'No Exp'
				data = data.decode('utf-8')
				data = trim(data)
				userdat.append(data)
				if driver.find_element_by_class_name('full-name').text != 'LinkedIn Member':
					alluserdat.append(userdat)
					userno += 1
				else:
					print 'Anonymous user. Not added to database'
			except NoSuchElementException:
				print 'element not found'
			print 'Done saving.'
		else:
			print 'Not a link :('
except:
    print 'Some exception'

with open('alluserdat', 'wt') as fout:
	csvout = csv.writer(fout)
	csvout.writerows(alluserdat)

#driver.quit()
