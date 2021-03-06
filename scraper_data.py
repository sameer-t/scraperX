#!/usr/bin/env python
# -*- coding: utf-8 -*-

from bs4 import BeautifulSoup as bs
import os
from openpyxl import Workbook

# A trim function to remove any special symbols and multiple linebreaks
def trim (des):
    desc = des.encode('utf-8',"ignore")
    toremove = ['•','\n–\n','\n–','\n1.','\n2.','\n3.','\n4.','\n5.','\n6.','\n7.','\n8.','\n9.','\n10.','\n11.','\n12.', '\n13.', '\n14.', '\n15', '\n\n', '\n\n\n']
    for symbol in toremove:
        if symbol in desc:
            desc = desc.replace(symbol,' ') 
    return desc

wb = Workbook()
ws = wb.active

# Move all the html files saved from scraper_html script into a folder and then run the following code
len_users = len(os.listdir('/home/sameer/linkedin-scraper/uber_profiles'))
for userno in range(1,len_users+1):
    profile = open('/home/sameer/linkedin-scraper/uber_profiles/'+ str(userno) + '.html', 'r')
    soup = bs(profile)
    try:
        name = trim(soup.find("span","full-name").string)
    except:
        name = 'No Name'
    try:
        title = trim(soup.find("p","title").text)
    except:
        title = 'No Title'
    try:
        loc = trim(soup.find("span","locality").text)
    except:
        loc = 'No Location'
    try:
        pub_profile = soup.find("dl","public-profile").dd.string
        if pub_profile == '':
            pub_profile = soup.find("div","full-top-card-unsaved-public-profile").text
        pub_profile = trim(pub_profile)
    except:
        pub_profile = 'No Public Profile'
    try:
        summary = trim(soup.find("div","summary").text)
    except:
        summary = 'No Summary'
    try:
        experience = ''
        exp = soup.find(id="background-experience")
        explist = exp.find_all('div','editable-item section-item')
        for expno in range(len(explist)):
            for string in explist[expno].stripped_strings:
                experience = experience + string + '\n'
        experience = trim(experience)
    except:
        experience = 'No Experience'
    try:
        skillset = ''
        skills = soup.find_all("span","endorse-item-name")
        for skill in skills:
            skillset = skillset + trim(skill.text) + '\n'
        skillset = trim(skillset)
    except:
        skillset = 'No Skills'
    try:
        education = ''
        edu = soup.find(id="background-education")
        edulist = edu.find_all('div','editable-item section-item')
        for eduno in range(len(edulist)):
            for string in edulist[eduno].stripped_strings:
                education = education + string + '\n'
        education = trim(education)
        education = education.replace('\n,\n',',')
    except:
        education = 'No education'
    try:
        interests = trim(soup.find("ul","interests-listing").text)
    except:
        interests = 'No interests'
    ws.cell(row = userno, column = 1).value = name.decode('utf-8','ignore')
    ws.cell(row = userno, column = 2).value = title.decode('utf-8','ignore')
    ws.cell(row = userno, column = 3).value = pub_profile.decode('utf-8','ignore')
    ws.cell(row = userno, column = 4).value = summary.decode('utf-8','ignore')
    ws.cell(row = userno, column = 5).value = experience.decode('utf-8','ignore')
    ws.cell(row = userno, column = 6).value = skillset.decode('utf-8','ignore')
    ws.cell(row = userno, column = 7).value = education.decode('utf-8','ignore')
    ws.cell(row = userno, column = 8).value = interests.decode('utf-8','ignore')
    ws.cell(row = userno, column = 9).value = loc.decode('utf-8','ignore')
    data = title + '\n' + loc + '\n' + summary + '\n' + experience + ' ' + skillset + ' ' + education + ' ' + interests
    data = trim(data)
    ws.cell(row = userno, column = 10).value = data.decode('utf-8','ignore')
    print str(userno) + ' ' + name + ' done'

# Save the spreadsheet. It is to be used for further formatting in Excel
wb.save('uber_raw.xlsx')
print 'Data scraping completed ( ͡° ͜ʖ ͡°)'