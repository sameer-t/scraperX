#!/usr/bin/env python
# -*- coding: utf-8 -*-

from copy import copy
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
import os
import time

def dur_mon (totdur):
    if 'year' in totdur:
        ny = int(totdur[totdur.find('(')+1:totdur.find('year')-1])
    else:
        ny = 0
    if 'month' in totdur and not 'year' in totdur:
        nm = int(totdur[totdur.find('(')+1:totdur.find('month')-1])
    elif 'month' in totdur and 'year' in totdur:
        nm = int(totdur[totdur.find('month')-3:totdur.find('month')-1])
    else:
        nm = 0
    tm = ny*12 + nm
    return tm

def trim (des):
    desc = des.encode('utf8')
    toremove = ['•', '-','\n1.','\n2.','\n3.','\n4.','\n5.','\n6.','\n7.','\n8.','\n9.','\n10.','\n11.','\n12.', '\n13.', '\n14.', '\n15']
    for symbol in toremove:
        if symbol in desc:
            desc = desc.replace(symbol,'') 
    return desc

def maploc (dop, place):
    for k in dop:
        for v in dop[k]:
            if place in v:
                return k
    return place

ctm = {}
maplocs = load_workbook(filename='/home/sameer/linkedin-scraper/cities to map.xlsx',use_iterators=True)
for area in maplocs.get_sheet_names():
    ctm[area] = []
    ts = maplocs[area]
    for row in ts.iter_rows():
        for cell in row:
            ctm[area].append(cell.value)

# Load driver.


wb = Workbook()
ws = wb.active
#wb.remove_sheet(ws)
#ws = wb.create_sheet()

len_users = len(os.listdir('/home/sameer/linkedin-scraper/NEST_profiles'))
userno = 1 
i = 1
rc0=0
rc1=0
rc2=0
rc3=0
try:
    for profno in range(1,len_users+1):
        driver = webdriver.Firefox()
        driver.get('file:///home/sameer/linkedin-scraper/NEST_profiles/'+ str(profno) + '.html')
        time.sleep(5)
        if 1:
            print userno,
            try:
                #Name
                try:
                    ws.cell(row = i, column = 1).value = driver.find_element_by_class_name('full-name').text
                    print ws.cell(row = i, column = 1).value
                except:
                    ws.cell(row = i, column = 1).value = 'No Name'
                    print 'No Name'            
                #Location
                try:
                    wrapper6 = driver.find_element_by_id('demographics')
                    ws.cell(row = i, column = 19).value = maploc (ctm,wrapper6.find_element_by_class_name('locality').text)
                except:
                    ws.cell(row = i, column = 19).value = 'Location Missing'
                #Public Profile
                try:
                    pp = driver.find_element_by_class_name('public-profile').text
                    if pp == '':
                        pp = driver.find_element_by_class_name('full-top-card-unsaved-public-profile').text
                    if pp != '':
                        ws.cell(row = i, column = 20).value = pp
                    else:
                        ws.cell(row = i, column = 20).value = 'Public Profile absent'
                        print 'Public Profile absent'
                except:
                    ws.cell(row = i, column = 20).value = 'Public Profile absent'
                    print 'Public Profile absent'
                #Title
                try:
                    ws.cell(row = i, column = 2).value = driver.find_element_by_class_name('title').text
                    print 'Title Done'
                except:
                    ws.cell(row = i, column = 2).value = 'No Title'
                    print 'No Title' 
                #Summary                        
                try:
                    wrapper1 = driver.find_element_by_class_name('summary')
                    ws.cell(row = i, column = 3).value = wrapper1.find_element_by_class_name('description').text
                    ws.cell(row = i, column = 18).value = wrapper1.find_element_by_class_name('description').text
                except NoSuchElementException:
                    ws.cell(row = i, column = 3).value = 'Not listed'
                    ws.cell(row = i, column = 18).value = ''
                print 'Summary Done'
                #Skills
                try:
                    rc0 = copy(i)
                    wrapper3 = driver.find_elements_by_class_name('endorse-item-name')
                    for v in range(len(wrapper3)):
                        ski = wrapper3[v].text.replace('...','')
                        if ski != '':
                            ws.cell(row = rc0, column = 4).value = ski
                            rc0 += 1
                    print 'Skills Done'
                except:
                    ws.cell(row = i, column = 4).value = 'No Skills'
                    print 'No Skills'
                #Education
                try:
                    rc1 = copy(i)
                    wrapper4 = driver.find_element_by_id('background-education')
                    er = wrapper4.find_elements_by_class_name('editable-item')
                    for l in range(len(er)):
                        try:
                            ws.cell(row = rc1, column = 5).value = er[l].find_element_by_class_name('summary').text
                        except:
                            print 'School name missing'
                        try:
                            ws.cell(row = rc1, column = 6).value = er[l].find_element_by_class_name('degree').text.replace(',','')
                        except:
                            print 'Degree missing'
                        try:
                            ws.cell(row = rc1, column = 7).value = er[l].find_element_by_class_name('major').text
                        except:
                            print 'Major missing'
                        try:
                            ws.cell(row = rc1, column = 8).value = er[l].find_element_by_class_name('education-date').text
                        except:
                            print 'Date missing'
                        try:
                            ws.cell(row = rc1, column = 9).value = trim(er[l].find_element_by_class_name('activities').text[26:])
                        except:
                            print 'Activities missing'
                        rc1 +=1
                    print 'Edu Done'
                except:
                    ws.cell(row = i, column = 5).value = 'No Education'
                    print 'No Education'
                #Additional Info - Interests
                try:
                    rc3 = copy(i)
                    wrapper5 = driver.find_element_by_class_name('interests-listing')
                    for userint in wrapper5.text.split(', '):
                        ws.cell(row = rc3, column = 10).value = userint
                        rc3 += 1
                except NoSuchElementException:
                    ws.cell(row = i, column = 10).value = None
                print 'Int Done'
                #Experience
                try:
                    rc2 = copy(i)
                    wrapper2 = driver.find_element_by_id('background-experience')
                    qw = wrapper2.find_elements_by_class_name('editable-item')
                    for k in range(len(qw)):
                        try:
                            temp1 = qw[k].find_element_by_tag_name('header').text.split('\n')
                            ws.cell(row = rc2, column = 11).value = temp1[0]
                            ws.cell(row = rc2, column = 12).value = temp1[1]
                        except:
                            print 'Something in header missing'    
                        try:    
                            temp2 = qw[k].find_elements_by_tag_name('time')
                            ws.cell(row = rc2, column = 13).value = temp2[0].text
                            try:
                                ws.cell(row = rc2, column = 14).value = temp2[1].text
                            except IndexError:
                                ws.cell(row = rc2, column = 14).value = 'Present'
                            ws.cell(row = rc2, column = 15).value = dur_mon(qw[k].find_element_by_class_name('experience-date-locale').text)
                        except:
                            print 'Something in time missing'
                        try:
                            ws.cell(row = rc2, column = 16).value = maploc (ctm,qw[k].find_element_by_class_name('locality').text)
                        except:
                            print 'location of exp missing'
                        try:
                            ws.cell(row = rc2, column = 17).value = trim(qw[k].find_element_by_class_name('description').text)
                            ws.cell(row = i, column = 18).value = ws.cell(row = i, column = 18).value+' '+ws.cell(row = rc2, column = 17).value
                        except:
                            print 'No description of exp'
                        rc2 += 1
                    print 'Exp Done'
                except:
                    ws.cell(row = rc2, column = 11).value = 'No Experience'
                    print 'No Exp'
            except NoSuchElementException:
                print 'element not found'
            print 'Done saving.'
            userno += 1
            i = max(rc0,rc1,rc2,rc3,i+1)
        else:
            print 'Not a link :('
        driver.quit()
except:
    print 'Some exception'

wb.save('NEST_test2.xlsx')